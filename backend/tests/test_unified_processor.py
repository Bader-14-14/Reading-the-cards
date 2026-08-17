from fastapi.testclient import TestClient

from app.main import app
from app.unified_processor import classify_card_text


CLIENT = TestClient(app)


def test_card_classification_uses_content_not_filename():
    assert classify_card_text("هوية مقيم رقم الهوية").card_type == "iqama"
    assert classify_card_text("مقيم هوية رقم الهوية الجنسية").card_type == "iqama"
    assert classify_card_text("رخصة قيادة License").card_type == "driving_license"
    assert classify_card_text("استمارة رخصة سير").card_type == "vehicle_registration"
    assert classify_card_text("الهوية الوطنية National ID").card_type == "national_id"


def test_license_parser_fields_are_independent_from_identity_tables():
    from app.parsers import parse_license

    parsed = parse_license(
        "رخصة قيادة\n"
        "MOHAMMAD SALEEM MOHAMMAD NASEEM\n"
        "ID Number: 2572312086\n"
        "License Type: Heavy Transport\n"
        "Issue Date: 05/03/2011\n"
        "Date of Birth: 08/08/1988\n"
        "Nationality: India\n"
        "Expiry Date: 31/07/2030\n"
        "Blood Type: B+",
        language="en",
    )

    assert parsed["license_number"] == "2572312086"
    assert parsed["license_type"] == "Heavy Transport"
    assert parsed["issue_date"] == "05/03/2011"
    assert parsed["blood_type"] == "B+"


def test_license_parser_selects_arabic_values_when_requested():
    from app.parsers import parse_license

    parsed = parse_license(
        "محمد سليم محمد نسيم MOHAMMAD SALEEM MOHAMMAD NASEEM\n"
        "نوع الرخصة: نقل ثقيل\nLicense Type: Heavy Transport\n"
        "الهند الجنسية\nNationality: India\nفصيلة الدم: +B",
        language="ar",
    )

    assert parsed["name"] == "محمد سليم محمد نسيم"
    assert parsed["license_type"] == "نقل ثقيل"
    assert parsed["nationality"] == "الهند"


def test_license_export_order_matches_card_order(tmp_path):
    from openpyxl import load_workbook
    from app.exporter import create_excel

    output = tmp_path / "license.xlsx"
    create_excel(
        {
            "name": "محمد سليم محمد نسيم",
            "id_number": "2572312086",
            "license_number": "2572312086",
            "license_type": "نقل ثقيل",
            "issue_date": "2011/03/05",
            "dob": "1988/08/08",
            "nationality": "الهند",
            "expiry": "2030/07/31",
            "blood_type": "+B",
        },
        str(output),
        language="ar",
    )

    rows = list(load_workbook(output).active.iter_rows(values_only=True))
    assert [row[0] for row in rows] == [
        "الحقل",
        "نوع البطاقة",
        "الاسم",
        "رقم الهوية",
        "رقم الرخصة",
        "نوع الرخصة",
        "تاريخ الإصدار",
        "تاريخ الميلاد",
        "الجنسية",
        "تاريخ الانتهاء",
        "فصيلة الدم",
    ]


def test_license_fields_stop_at_adjacent_labels_and_spaces():
    from app.parsers import parse_license

    parsed = parse_license(
        "نوع الرخصة: خصوصي تاريخ الإصدار: ٢٠٢٣/١٠/٢٤ "
        "تاريخ الميلاد: ١٩٩٣/٠١/٠٣ باكستان الجنسية: "
        "تاريخ الانتهاء: ٢٠٢٨/٠٨/٣٠ فصيلة الدم: +B",
        language="ar",
    )

    assert parsed["license_type"] == "خصوصي"
    assert parsed["nationality"] == "باكستان"
    assert parsed["blood_type"] == "+B"


def test_unified_endpoint_accepts_one_or_many_and_keeps_errors_per_file(monkeypatch):
    calls = []

    def fake_process(image_bytes, *, provider, language):
        calls.append(image_bytes)
        if image_bytes == b"bad":
            raise ValueError("bad image")
        return {
            "card_type": "iqama",
            "classification_confidence": "high",
            "language": language,
            "data": {"name": "test", "id_number": "2422262553"},
            "missing_fields": ["nationality"],
        }

    monkeypatch.setattr("app.main.process_card", fake_process)
    response = CLIENT.post(
        "/extract-cards?language=en",
        files=[
            ("files", ("random-name.bin", b"good", "image/jpeg")),
            ("files", ("another-name", b"bad", "image/jpeg")),
        ],
    )

    assert response.status_code == 200
    results = response.json()["results"]
    assert len(results) == 2
    assert results[0]["card_type"] == "iqama"
    assert results[0]["filename"] == "random-name.bin"
    assert results[1]["error"] == "Card processing failed."
    assert calls == [b"good", b"bad"]
