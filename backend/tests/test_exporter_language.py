from openpyxl import load_workbook

from app.exporter import create_excel


def test_english_export_uses_english_labels_and_values(tmp_path):
    output = tmp_path / 'iqama.xlsx'
    create_excel(
        {
            'name': 'MOHAMMAD SALEEM MOHAMMAD NASEEM',
            'iqama_number': '2572312086',
            'nationality': 'India',
            'nationality_ar': 'الهند',
            'dob': '1988/08/08',
            'doe': '2026/10/06',
            'raw_text': 'internal OCR',
        },
        str(output),
        language='en',
    )

    rows = list(load_workbook(output).active.iter_rows(values_only=True))
    assert rows == [
        ('Field', 'Value'),
        ('Card Type', 'Iqama'),
        ('Name', 'MOHAMMAD SALEEM MOHAMMAD NASEEM'),
        ('Iqama Number', '2572312086'),
        ('Nationality', 'India'),
        ('Date of Birth', '1988/08/08'),
        ('Date of Expiry', '2026/10/06'),
    ]


def test_license_export_starts_with_card_type_and_matches_card_order(tmp_path):
    output = tmp_path / 'license.xlsx'
    create_excel(
        {
            'name': 'MOHAMMAD SALEEM MOHAMMAD NASEEM',
            'id_number': '2572312086',
            'license_number': '2572312086',
            'license_type': 'Heavy Transport',
            'issue_date': '2011/03/05',
            'dob': '1988/08/08',
            'nationality': 'India',
            'expiry': '2030/07/31',
            'blood_type': 'B+',
        },
        str(output),
        language='en',
    )

    rows = list(load_workbook(output).active.iter_rows(values_only=True))
    assert rows == [
        ('Field', 'Value'),
        ('Card Type', 'Driving License'),
        ('Name', 'MOHAMMAD SALEEM MOHAMMAD NASEEM'),
        ('ID Number', '2572312086'),
        ('License Type', 'Heavy Transport'),
        ('Issue Date', '2011/03/05'),
        ('Date of Birth', '1988/08/08'),
        ('Nationality', 'India'),
        ('Date of Expiry', '2030/07/31'),
        ('Blood Type', 'B+'),
    ]